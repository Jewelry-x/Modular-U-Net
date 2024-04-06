# This is part of the tutorial materials in the UCL Module MPHY0041: Machine Learning in Medical Imaging
import os
import numpy as np
import winsound

from openpyxl import Workbook

import torch
from torch.utils.data import DataLoader, Dataset
from torchvision.transforms import v2

from config import *

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
use_cuda = torch.cuda.is_available()

augmentations = []

if HORIZONTAL_FLIP:
    augmentations.append("\nHorizontal Flip")
if VERTICAL_FLIP:
    augmentations.append("\nVertical Flip")
if SHEAR:
    augmentations.append("\nShear")
if GAUSSIAN_BLUR:
    augmentations.append("\nGaussian Blur")


## network class
class UNet(torch.nn.Module):
    def __init__(self, ch_in=1, ch_out=1, init_n_feat=32):
        super(UNet, self).__init__()

        n_feat = init_n_feat
        self.encoder1 = UNet._block(ch_in, n_feat)
        self.pool1 = torch.nn.MaxPool2d(kernel_size=2, stride=2)
        self.encoder2 = UNet._block(n_feat, n_feat * 2)
        self.pool2 = torch.nn.MaxPool2d(kernel_size=2, stride=2)
        self.encoder3 = UNet._block(n_feat * 2, n_feat * 4)
        self.pool3 = torch.nn.MaxPool2d(kernel_size=2, stride=2)
        self.encoder4 = UNet._block(n_feat * 4, n_feat * 8)
        self.pool4 = torch.nn.MaxPool2d(kernel_size=2, stride=2)

        self.bottleneck = UNet._block(n_feat * 8, n_feat * 16)

        self.upconv4 = torch.nn.ConvTranspose2d(
            n_feat * 16, n_feat * 8, kernel_size=2, stride=2
        )
        self.decoder4 = UNet._block((n_feat * 8) * 2, n_feat * 8)
        self.upconv3 = torch.nn.ConvTranspose2d(
            n_feat * 8, n_feat * 4, kernel_size=2, stride=2
        )
        self.decoder3 = UNet._block((n_feat * 4) * 2, n_feat * 4)
        self.upconv2 = torch.nn.ConvTranspose2d(
            n_feat * 4, n_feat * 2, kernel_size=2, stride=2
        )
        self.decoder2 = UNet._block((n_feat * 2) * 2, n_feat * 2)
        self.upconv1 = torch.nn.ConvTranspose2d(
            n_feat * 2, n_feat, kernel_size=2, stride=2
        )
        self.decoder1 = UNet._block(n_feat * 2, n_feat)

        self.conv = torch.nn.Conv2d(
            in_channels=n_feat, out_channels=ch_out, kernel_size=1
        )

        if REVERSE_POOL:
            if POOL < 4:
                self.decoder1 = UNet._block(n_feat * 3, n_feat)
                if POOL < 3:
                    self.decoder2 = UNet._block((n_feat * 2) * 3, n_feat * 2)
                    if POOL < 2:
                        self.decoder3 = UNet._block((n_feat * 4) * 3, n_feat * 4)
                        if POOL < 1:
                            self.bottleneck = UNet._block(n_feat * 8, n_feat * 8)
        else:
            if POOL < 4:
                self.bottleneck = UNet._block(n_feat * 8, n_feat * 8)
                if POOL < 3:
                    self.decoder3 = UNet._block((n_feat * 4) * 3, n_feat * 4)
                    if POOL < 2:
                        self.decoder2 = UNet._block((n_feat * 2) * 3, n_feat * 2)
                        if POOL < 1:
                            self.decoder1 = UNet._block(n_feat * 3, n_feat)

    def forward(self, x):
        enc1 = self.encoder1(x)
        enc2 = self.encoder2(
            enc1
            if POOL < 4 and REVERSE_POOL == True or POOL < 1 and REVERSE_POOL == False
            else self.pool1(enc1)
        )
        enc3 = self.encoder3(
            enc2
            if POOL < 3 and REVERSE_POOL == True or POOL < 2 and REVERSE_POOL == False
            else self.pool1(enc2)
        )
        enc4 = self.encoder4(
            enc3
            if POOL < 2 and REVERSE_POOL == True or POOL < 3 and REVERSE_POOL == False
            else self.pool1(enc3)
        )

        bottleneck = self.bottleneck(
            enc4
            if POOL < 1 and REVERSE_POOL == True or POOL < 4 and REVERSE_POOL == False
            else self.pool1(enc4)
        )

        dec4 = (
            bottleneck
            if POOL < 1 and REVERSE_POOL == True or POOL < 4 and REVERSE_POOL == False
            else self.upconv4(bottleneck)
        )
        dec4 = torch.cat((dec4, enc4), dim=1)
        dec4 = self.decoder4(dec4)
        dec3 = (
            dec4
            if POOL < 2 and REVERSE_POOL == True or POOL < 3 and REVERSE_POOL == False
            else self.upconv3(dec4)
        )
        dec3 = torch.cat((dec3, enc3), dim=1)
        dec3 = self.decoder3(dec3)
        dec2 = (
            dec3
            if POOL < 3 and REVERSE_POOL == True or POOL < 2 and REVERSE_POOL == False
            else self.upconv2(dec3)
        )
        dec2 = torch.cat((dec2, enc2), dim=1)
        dec2 = self.decoder2(dec2)
        dec1 = (
            dec2
            if POOL < 4 and REVERSE_POOL == True or POOL < 1 and REVERSE_POOL == False
            else self.upconv1(dec2)
        )
        dec1 = torch.cat((dec1, enc1), dim=1)
        dec1 = self.decoder1(dec1)
        return torch.sigmoid(self.conv(dec1))

    @staticmethod
    def _block(ch_in, n_feat):
        return torch.nn.Sequential(
            torch.nn.Conv2d(
                in_channels=ch_in,
                out_channels=n_feat,
                kernel_size=3,
                padding=1,
                bias=False,
            ),
            torch.nn.BatchNorm2d(num_features=n_feat),
            torch.nn.ReLU(inplace=True),
            torch.nn.Conv2d(
                in_channels=n_feat,
                out_channels=n_feat,
                kernel_size=3,
                padding=1,
                bias=False,
            ),
            torch.nn.BatchNorm2d(num_features=n_feat),
            torch.nn.ReLU(inplace=True),
        )


def calculate_2d_metrics(gt_slice, pred_slice):
    intersection = np.logical_and(gt_slice, pred_slice)
    union = np.logical_or(gt_slice, pred_slice)

    iou = np.sum(intersection) / np.sum(union)
    dice_coefficient = (
        2.0 * np.sum(intersection) / (np.sum(gt_slice) + np.sum(pred_slice))
    )

    return iou, dice_coefficient


def calculate_average_metrics(gt_volume, pred_volume):
    num_slices = gt_volume.shape[0]
    iou_scores = []
    dice_scores = []

    for slice_idx in range(num_slices):
        gt_slice = gt_volume[slice_idx]
        skip = np.sum(gt_slice)
        if skip > 0:
            pred_slice = pred_volume[slice_idx]
            iou, dice_coefficient = calculate_2d_metrics(gt_slice, pred_slice)
            iou_scores.append(iou)
            dice_scores.append(dice_coefficient)

    avg_iou = np.mean(iou_scores)
    avg_dice = np.mean(dice_scores)

    return avg_iou, avg_dice


## loss function
def loss_2Ddice(y_pred, y_true, eps=1e-6):
    """
    y_pred, y_true -> [N, C=1, D, H, W]
    """
    numerator = torch.sum(y_true * y_pred, dim=(2, 3)) * 2
    denominator = torch.sum(y_true, dim=(2, 3)) + torch.sum(y_pred, dim=(2, 3)) + eps
    return torch.mean(1.0 - (numerator / denominator))


def update_excel(
    epoch, epoch_train_loss, epoch_val_loss, val_iou, model_saved, worksheet
):
    # Update values in respective rows and columns
    worksheet.cell(row=epoch + 1, column=1, value=f"Epoch {epoch}")
    worksheet.cell(row=epoch + 1, column=2, value=epoch_train_loss)
    worksheet.cell(row=epoch + 1, column=3, value=epoch_val_loss)
    worksheet.cell(row=epoch + 1, column=4, value=val_iou)

    if model_saved:
        worksheet.cell(row=epoch + 1, column=5, value="Saved")
    else:
        worksheet.cell(row=epoch + 1, column=5, value="")


## data loader
class NPyDataset(Dataset):
    def __init__(self, folder_name, is_train=True):
        self.folder_name = folder_name
        self.is_train = is_train
        self.transform = self._get_transform()

        global TRAIN_IMAGE_SIZE

        if TRAIN_IMAGE_SIZE == 0:
            image = self._load_npy(
                os.path.join(
                    TRAINING_DATA_LOCATION[TRAINING_DATA.index(DATA)],
                    IMAGE_DEFINITION % 0000,
                )
            )

            TRAIN_IMAGE_SIZE = image.shape[1]

        global TEST_IMAGE_SIZE

        if not is_train:
            if TEST_IMAGE_SIZE == 0:
                image = self._load_npy(
                    os.path.join(
                        TRAINING_DATA_LOCATION[TRAINING_DATA.index(DATA)],
                        IMAGE_DEFINITION % 0000,
                    )
                )

                TEST_IMAGE_SIZE = image.shape[1]

    def __len__(self):
        return (
            TRAINING_DATA_COUNT[TRAINING_DATA.index(DATA)]
            if self.is_train
            else TESTING_DATA_COUNT[TESTING_DATA.index(DATA)]
        )

    def __getitem__(self, idx):
        if self.is_train:
            image = self._load_npy(
                os.path.join(
                    TRAINING_DATA_LOCATION[TRAINING_DATA.index(DATA)],
                    IMAGE_DEFINITION % idx,
                )
            )
            label = self._load_npy(
                os.path.join(
                    TRAINING_DATA_MASK_LOCATION[TRAINING_DATA.index(DATA)],
                    MASK_DEFINITION % idx,
                )
            )
        else:

            image = self._load_npy(
                os.path.join(
                    TESTING_DATA_LOCATION[TESTING_DATA.index(DATA)],
                    IMAGE_DEFINITION % idx,
                )
            )
            label = self._load_npy(
                os.path.join(
                    TESTING_DATA_MASK_LOCATION[TESTING_DATA.index(DATA)],
                    MASK_DEFINITION % idx,
                )
            )

        if self.transform and TRANSFORM:
            image, label = self.transform(image, label)

        return image, label

    def _load_npy(self, filename):
        filename = os.path.join(self.folder_name, filename)
        if self.is_train:
            if TRAIN_REDUCE_SIZE == 1:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename))), dim=0
                )
            if TRAIN_REDUCE_SIZE == 2:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename)[::2, ::2])), dim=0
                )
            if TRAIN_REDUCE_SIZE == 4:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename)[::4, ::4])), dim=0
                )
        else:
            if TEST_REDUCE_SIZE == 1:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename))), dim=0
                )
            if TEST_REDUCE_SIZE == 2:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename)[::2, ::2])), dim=0
                )
            if TEST_REDUCE_SIZE == 4:
                return torch.unsqueeze(
                    torch.tensor(np.float32(np.load(filename)[::4, ::4])), dim=0
                )

    def _get_transform(self):
        if self.is_train:
            transformations = []

            if HORIZONTAL_FLIP:
                transformations.append(v2.RandomHorizontalFlip(p=0.5))
            if VERTICAL_FLIP:
                transformations.append(v2.RandomVerticalFlip(p=0.5))
            if SHEAR:
                transformations.append(v2.RandomAffine(0, shear=15))
            if GAUSSIAN_BLUR:
                transformations.append(v2.GaussianBlur(kernel_size=3))

            if transformations:
                return v2.Compose(transformations)
        return None


def train_load_data():
    # training data loader
    train_set = NPyDataset(DATA_PATH)
    train_loader = DataLoader(train_set, batch_size=4, shuffle=True, num_workers=0)

    return train_loader


def val_load_data():
    # validation data loader
    val_set = NPyDataset(DATA_PATH, is_train=False)
    val_loader = DataLoader(val_set, batch_size=1, shuffle=False, num_workers=0)

    return val_loader


def test_load_data():
    # test data loader
    test_set = NPyDataset(DATA_PATH, is_train=False)
    test_loader = DataLoader(test_set, batch_size=1, shuffle=False, num_workers=0)

    return test_loader


## training
def train(load=False):
    train_loader = train_load_data()
    val_loader = val_load_data()

    if SET_SEED:
        torch.manual_seed(SEED)
        torch.cuda.manual_seed(SEED)
        torch.cuda.manual_seed_all(SEED)  # if you are using multiple GPUs
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False

    if not load:
        print("Starting new model")
        model = UNet(
            1, 1
        )  # input 1-channel 3d volume and output 1-channel segmentation (a probability map)
        if use_cuda:
            model.cuda()

        best_eval_loss = 9999
    else:
        print("Loading model")

        if not os.path.exists(os.path.join(RESULT_PATH, "saved_model_pt")):
            print("Model not in correct directory (.result/)")
            exit()
        whole_model = torch.load(os.path.join(RESULT_PATH, "saved_model_pt"))
        print_model_values(whole_model)

        if "model" in whole_model:
            model = whole_model["model"]
        else:
            model = whole_model

        if "pools" in whole_model:
            print("Pooling layers used: " + str(whole_model.get("pools")))

            global POOL
            POOL = whole_model.get("pools")

        y_pred_test = np.array([])
        ytest = np.array([])
        for _, (images, labels) in enumerate(val_loader):

            images = images.cuda()
            output = model(images)
            predicted = np.float32(
                np.squeeze(output.detach().cpu().numpy(), axis=0) > 0.5
            )
            gt = np.squeeze(labels.detach().cpu().numpy(), axis=0)
            if y_pred_test.size == 0:
                y_pred_test = predicted
                ytest = gt
            else:
                y_pred_test = np.concatenate((y_pred_test, predicted), axis=0)
                ytest = np.concatenate((ytest, gt), axis=0)

        _, dice = calculate_average_metrics(ytest, y_pred_test)

        best_eval_loss = 1 - dice

    num_epochs = int(TOTAL_EPOCHS)
    optimizer = torch.optim.AdamW(model.parameters(), lr=LEARNING_RATE)

    os.makedirs("result", exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Epoch", "Train Loss", "Val Loss", "Val IoU", "Model Saved"])

    global augmentations
    if augmentations:
        augmentations = ", ".join(augmentations)
    else:
        augmentations = ""

    no_loss_change = 0

    for epoch in range(num_epochs):
        no_loss_change += 1

        running_train_loss = 0.0
        correct_train = 0
        total_train = 0
        train_iou = 0.0

        for _, (images, labels) in enumerate(train_loader):
            if use_cuda:
                images, labels = images.cuda(), labels.cuda()

            optimizer.zero_grad()
            outputs = model(images)

            loss = loss_2Ddice(outputs, labels)

            loss.backward()
            optimizer.step()
            running_train_loss += loss.item()
            iou, _ = calculate_average_metrics(
                np.squeeze(outputs.detach().cpu().numpy()),
                np.squeeze(labels.detach().cpu().numpy()),
            )
            pred_flat = outputs.view(-1) > 0.5
            target_flat = labels.view(-1)
            total_train += target_flat.size(0)
            train_iou += iou
            correct_train += (pred_flat == target_flat).sum().item()

        ## validation
        y_pred_test = np.array([])
        ytest = np.array([])
        for _, (images, labels) in enumerate(val_loader):

            images = images.cuda()
            output = model(images)
            predicted = np.float32(
                np.squeeze(output.detach().cpu().numpy(), axis=0) > 0.5
            )
            gt = np.squeeze(labels.detach().cpu().numpy(), axis=0)
            if y_pred_test.size == 0:
                y_pred_test = predicted
                ytest = gt
            else:
                y_pred_test = np.concatenate((y_pred_test, predicted), axis=0)
                ytest = np.concatenate((ytest, gt), axis=0)

        val_iou, dice = calculate_average_metrics(ytest, y_pred_test)
        train_iou = train_iou / len(train_loader)
        epoch_train_loss = running_train_loss / len(train_loader)
        epoch_val_loss = 1 - dice

        print(
            f"Epoch {epoch+1}, Train Loss: {epoch_train_loss:.4f}, Validation Loss: {epoch_val_loss:.4f}, Val iou: {val_iou:.2f}"
        )

        model_saved = False

        if epoch_val_loss < best_eval_loss:
            no_loss_change = 0
            model_saved = True
            best_eval_loss = epoch_val_loss
            torch.save(
                {
                    "model": model,
                    "data": DATA,
                    "learning_rate": LEARNING_RATE,
                    "pools": POOL,
                    "reverse_pools": str(REVERSE_POOL),
                    "data_augmentations": augmentations,
                    "image_size": TRAIN_IMAGE_SIZE,
                    "optimizer": type(optimizer).__name__,
                    "epoch": epoch + 1,
                    "early_stopping": str(EARLY_STOPPING),
                    "early_stopping_epochs": str(EARLY_STOPPING_COUNT),
                },
                os.path.join(RESULT_PATH, "saved_model_pt"),
            )
            print("Model Saved Successfully")

            whole_model = torch.load(os.path.join(RESULT_PATH, "saved_model_pt"))
            print_model_values(whole_model)

        update_excel(
            epoch + 1, epoch_train_loss, epoch_val_loss, val_iou, model_saved, worksheet
        )

        workbook.save("./result/progress.xlsx")

        if no_loss_change == EARLY_STOPPING_COUNT and EARLY_STOPPING:
            break

    print("Training done.")


def test():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    if not os.path.exists(os.path.join(RESULT_PATH, "saved_model_pt")):
        print("Model not in correct directory (.result/)")
        exit()
    whole_model = torch.load(os.path.join(RESULT_PATH, "saved_model_pt"))
    print_model_values(whole_model)

    global POOL
    POOL = whole_model.get("pools")

    if "model" in whole_model:
        model = whole_model["model"]
    else:
        model = whole_model

    iou_arr, dc_arr, data_arr = [], [], []

    for data in TEST_ON:
        global DATA
        DATA = TESTING_DATA[data]
        data_arr.append(DATA)

        test_loader = test_load_data()
        print("\n\nData testing on: " + TESTING_DATA[data])
        print("Current image size: " + str(TEST_IMAGE_SIZE))


        model = model.to(device)
        model.eval()
        y_pred_test = np.array([])
        ytest = np.array([])
        iou = []
        dice_coefficient = []

        mask_create = CREATE_TEST_MASK
        for _, (images, labels) in enumerate(test_loader):
            images = images.cuda()
            output = model(images)
            predicted = np.float32(
                np.squeeze(output.detach().cpu().numpy(), axis=0) > 0.5
            )

            if mask_create:
                filepath_to_save = os.path.join(RESULT_PATH, DATA + "_test_label.npy")
                np.save(filepath_to_save, predicted)
                mask_create = False

            gt = np.squeeze(labels.detach().cpu().numpy(), axis=0)
            if y_pred_test.size == 0:
                y_pred_test = predicted
                ytest = gt
            else:
                y_pred_test = np.concatenate((y_pred_test, predicted), axis=0)
                ytest = np.concatenate((ytest, gt), axis=0)

            iou_val, dice_coefficient_val = calculate_2d_metrics(gt, predicted)

            iou.append(iou_val)
            dice_coefficient.append(dice_coefficient_val)

        iou_val, dice_coefficient_val = calculate_average_metrics(ytest, y_pred_test)
        print("Final IOU:", iou_val)
        print("Final DC:", dice_coefficient_val)

        iou_arr.append(iou_val)
        dc_arr.append(dice_coefficient_val)

    whole_model["tested_on"] = data_arr
    whole_model["IOU"] = iou_arr
    whole_model["DC"] = dc_arr

    torch.save(whole_model, os.path.join(RESULT_PATH, "saved_model_pt"))


def print_model_values(whole_model):
    if "data" in whole_model:
        print("Data trained on: " + str(whole_model.get("data")))
    if "learning_rate" in whole_model:
        print("Learning rate: " + str(whole_model.get("learning_rate")))
    if "pools" in whole_model:
        print("Pooling layers used: " + str(whole_model.get("pools")))
    if "reverse_pools" in whole_model:
        print(
            "Were the pooling layers removed in reverse: "
            + str(whole_model.get("reverse_pools"))
        )
    if "data_augmentations" in whole_model:
        print("Data Augmentations: " + str(whole_model.get("data_augmentations")))
    if "image_size" in whole_model:
        print("Training Image Size: " + str(whole_model.get("image_size")))
    if "optimizer" in whole_model:
        print("Optimizer: " + str(whole_model.get("optimizer")))
    if "epoch" in whole_model:
        print("Epoch Saved: " + str(whole_model.get("epoch")))
    if "early_stopping" in whole_model:
        print("Early Stopping: " + str(whole_model.get("early_stopping")))
    if "early_stopping_epochs" in whole_model:
        print(
            "Early Stopping Epoch Limit: "
            + str(whole_model.get("early_stopping_epochs"))
        )


def create_folder():
    whole_model = torch.load(os.path.join(RESULT_PATH, "saved_model_pt"))
    if (
        "optimizer" in whole_model
        and "data" in whole_model
        and "image_size" in whole_model
        and "learning_rate" in whole_model
        and "pools" in whole_model
        and "reverse_pools" in whole_model
        and "data_augmentations" in whole_model
    ):
        data_augmentations = whole_model.get("data_augmentations")
        if data_augmentations == "":
            data_augmentations = "None"

        data_augmentations = data_augmentations.replace("\n", "")

        folder_path = os.path.join(
            "models",
            str(whole_model.get("data")),
            str(whole_model.get("optimizer")),
            str(whole_model.get("image_size")),
            str(whole_model.get("learning_rate")),
            str(whole_model.get("pools")),
            str(whole_model.get("reverse_pools")),
            data_augmentations,
        )
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
    else:
        print("Model too outdated to create folder automatically")


if __name__ == "__main__":
    if not os.path.exists("result"):
        os.makedirs("result")

    if TRAIN:
        if LOAD:
            train(True)
        else:
            train()

    if TEST:
        test()

    if CREATE_FOLDER:
        create_folder()

    if NOTIFY:
        winsound.Beep(1000, 500)
